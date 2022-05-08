import logging
import datetime
import shelve
import os
import sys
from tkinter import FALSE
import time
import traceback
from contextlib import suppress
import threading
import typing
from typing import List, Dict, Mapping, Tuple, Union, Optional
import random

import cv2
import numpy as np
from pyzbar.pyzbar import decode
import openpyxl
from openpyxl.styles import Font
import pyinputplus
import ezsheets
from ezsheets import getColumnLetterOf, getColumnNumberOf
from ezsheets import EZSheetsException
import smtplib
import httplib2
import pyinputplus as pyip
from google.auth.exceptions import TransportError
from ssl import SSLEOFError

from quotes import strippedQuotes

# logging.basicConfig(filename='log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logging.disable(logging.DEBUG)

logging.info(f'File System Encoding = {sys.getfilesystemencoding()}')

print(
    """QRCODE ATTENDANCE RECORDING AUTOMATION PROGRAM
Programmed and maintained by
Adrian Luke Labasan | G11-Oxygen | CONTACT: zionexodus7@protonmail.com ~ 09157694749\n
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
     ??? INSTRUCTIONS ???
* Place the excel file at the same directory where the program (qrcode_attendance.py) is located.
* If first used, the excel file should be completely blank i.e. no data added.
* Press Ctrl+C to end the program.
* If you would like to change the layout of the excel file, please notify the programmer.

     !!!   WARNING    !!!
* Don't manipulate the excel file while the program is running!
* Before running the program, ensure that:
   (1) The excel file is closed.
   (2) The excel file is not manipulated/changed in any way beforehand except by the program.
* DO NOT DELETE the backup.* files. These files are used in the backup system and in the sync program. Delete the files only if you want to use the program for another set of data i.e. new set of students/members, new schoolyear, etc. (It won't affect the xlsx file.)
* The googleSheetsConfig.* files contain your configuration for the online database (Google Sheets). Delete the files if you want to reset them.
* The emailNotifConfig.* files contain your configuration for the email notification function (Gmail). Delete the files if you want to reset them.
* The settings.* files also contain the settings of the behavior of the program.
* DO NOT DELETE the token files. They contain your sign-in configuration for Google Sheets and Gmail.
* If you want to delete the spreadsheet in Google Sheets (ex. if you want to reupload it), you have to delete it also in the trash folder in your google drive. If you don't, the program can still detect it and then use it as its online database.

     ***    NOTE      ***
* The program only records the initial time a qrcode is scanned during the day.
* The order of the names are based on the time the qrcodes are first scanned. If you would like to sort the names by their section (alphabetically), you can do so in Excel: (Select data range by row) -> 'Data' tab -> 'SORT'. In Google Sheets: (Select data range by row) -> 'Data' tab -> 'Sort range'.
* You can only add the members' email and phone number after the xlsx file is initialized (in the first run).
* The program reduces the row and column counts in the online spreadsheet to make the read/write process faster. The program will automatically increase the row and column counts as the data grows.

If you want to suggest additional functionalities or modifications to the program, feel free to contact the programmer :)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n"""
)


# Local database configuration interface
def getExcelFiles() -> List[str]:
    excelFiles = []
    for file in os.listdir('.'):
        if file.endswith('.xlsx'):
            excelFiles.append(file)
    return excelFiles

while True:
    excelFiles = getExcelFiles()

    noInitialFile = False

    if len(excelFiles) == 0:
        print('No xlsx files found in the directory.')
        print('Please enter a filename for the excel file to be created. Press enter if you would like the default name [GARSHS_ATTENDANCE.xlsx]')
        excelFilename = pyinputplus.inputStr(prompt='> ', blank=True)
        if excelFilename == '':
            excelFilename = 'GARSHS_ATTENDANCE.xlsx'
        if not excelFilename.endswith('.xlsx'):
            excelFilename += '.xlsx'
        noInitialFile = True
        break
    elif len(excelFiles) == 1:
        excelFilename = excelFiles[0]
        print(f'Found {excelFilename}')
        print('Exit the program (Ctrl+C) if this is not the excel file you want.')
        break
    elif len(excelFiles) == 2 and (excelFiles[0].startswith('~$') or excelFiles[1].startswith('~$')):
        print('!!! The excel file is open in another window !!!')
        print('Please close the excel file then press enter to continue.')
        continueProgram = input('> ')
        continue
    else:
        exelFileIsOpen = False
        for file in os.listdir('.'):
            if exelFileIsOpen:
                break
            if file.startswith('~$') and file.endswith('.xlsx'):
                print('!!! An excel file is open in another window !!!')
                print('Please close the excel file/s then press enter to continue.')
                continueProgram = input('> ')
                for file in os.listdir('.'):
                    if file.startswith('~$') and file.endswith('.xlsx'):
                        exelFileIsOpen = True
                        break
        if exelFileIsOpen:
            continue
        print('Multiple xlsx files found in the directory.')
        excelFilename = pyinputplus.inputMenu(getExcelFiles(), numbered=True)
        break


# Online database configuration interface
def countdown(seconds=3):
    print(f'Retrying in {seconds} ', end='', flush=True)
    time.sleep(1)
    count = seconds - 1
    while count > 0:
        print(f'{count} ', end='', flush=True)
        count = count - 1
        time.sleep(1)
    print()


def silentCountdown(seconds=3):
    for second in range(seconds):
        time.sleep(1)


def credentialsSheetsInDirectory() -> bool:
    for filename in os.listdir('.'):
        if filename == 'credentials-sheets.json':
            return True
    return False


useOnlineDatabase = True
if not credentialsSheetsInDirectory():
    exitSetup = False
    useOnelineDatabaseResponse = pyip.inputYesNo('Do you want to use an online database (Google Sheets)? [yes/no]\n> ')
    while not exitSetup:
        if useOnelineDatabaseResponse == 'no':
            useOnlineDatabase = False
            break
        else:
            print("""Please follow these steps to be able to use Google Sheets:
1) Enable Google Sheets and Google Drive APIs through these links:
    * https://console.developers.google.com/apis/library/sheets.googleapis.com/
    * https://console.developers.google.com/apis/library/drive.googleapis.com/
2) Create and/or download credential file (OAuth client ID) through this link https://console.cloud.google.com/apis/credentials/oauthclient
    - You may have to configure the consent screen to create the credential file
3) Rename the downloaded file (json) to 'credentials-sheets.json' and place it in the same folder as the program
    - After you continue the program, you would have to sign in and authorize the app (twice)

NOTE: Always connect to the internet if you plan to use an online database.""")
            print('Press enter if done.')
            input('> ')
            exitSetup = credentialsSheetsInDirectory()

# Checks if the xlsx file is already uploaded to Google Sheets
# and if it is, gets its spreadsheetId
def spreadsheetExists():
    spreadsheets = ezsheets.listSpreadsheets()
    for key in spreadsheets:
        if spreadsheets[key] == excelFilename[:-5]:
            return True, key
    return False, None


if useOnlineDatabase:
    try:
        with shelve.open('googleSheetsConfig') as googleSheetsConfig:
            uploadedToGoogleSheets = googleSheetsConfig['uploadedToGoogleSheets']
    except KeyError:
        print('\nDetecting if the xlsx file is already uploaded to Google Sheets... ', end='', flush=True)
        uploadedToGoogleSheets = False
        while True:
            try:
                uploadedToGoogleSheets, spreadsheetId = spreadsheetExists()
                if uploadedToGoogleSheets:
                    with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                        googleSheetsConfig['uploadedToGoogleSheets'] = True
                        googleSheetsConfig['spreadsheetId'] = spreadsheetId
                    print('The spreadsheet is found in the server.')
                else:
                    with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                        googleSheetsConfig['uploadedToGoogleSheets'] = False
                    print('The spreadsheet is not found in the server. The workbook will be uploaded when the program ends.')
                break
            except (ConnectionAbortedError, httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, TransportError, SSLEOFError):
                with open('log.txt', 'a', encoding='utf-8') as log:
                    log.write(traceback.format_exc())
                print('Connection error. Please check your internet connection.')
                countdown()
                continue


class ExtendedSheet:
    def __init__(self, ss, title):
        self.sheet: ezsheets.Sheet = ss[title]

    
    def dateInColumnHeaders(self, date):
        row2 = self.sheet.getRow(2)
        for cell in row2:
            if cell == date:
                return True
        return False
    

    def nameInRowHeaders(self, name):
        col1 = self.sheet.getColumn(1)
        for cell in col1:
            if cell == name:
                return True
        return False


    def findFirstBlankCol(self):
        row2 = self.sheet.getRow(2)
        colNum = 1
        while row2[colNum - 1] != '':
            colNum += 1
        return colNum
    

    def findFirstBlankRow(self):
        col1 = self.sheet.getColumn(1)
        rowNum = 1
        while col1[rowNum - 1] != '':
            rowNum += 1
        return rowNum


if useOnlineDatabase and uploadedToGoogleSheets:
    ACTIVE_SHEET = 'Sheet'
    while True:
        with shelve.open('googleSheetsConfig') as googleSheetsConfig:
            spreadsheetId = googleSheetsConfig['spreadsheetId']
        try:
            print('\nConnecting to Google Sheets... ', end='', flush=True)
            ss = ezsheets.Spreadsheet(spreadsheetId)
            sheet1: ExtendedSheet = ExtendedSheet(ss, ACTIVE_SHEET)
            print('Connected.')
            break
        except (httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, \
            TransportError, SSLEOFError):
            with open('log.txt', 'a', encoding='utf-8') as log:
                log.write(traceback.format_exc())
            print('Failed to connect to Google Sheets. Please check your internet connection.')
            countdown()
            continue


# Email notification feature configuration interface
""" # Asks if the email notification function shall be used
try:
    with shelve.open('emailNotifConfig') as emailNotifConfig:
        useEmailNotification = emailNotifConfig['useEmailNotification']
except KeyError:
    useEmailNotificationResponse = pyip.inputYesNo('\nDo you want to use the email notification feature? [yes/no]\n> ')
    if useEmailNotificationResponse == 'no':
        with shelve.open('emailNotifConfig') as emailNotifConfig:
            emailNotifConfig['useEmailNotification'] = False
        useEmailNotification = False
    else:
        with shelve.open('emailNotifConfig') as emailNotifConfig:
            emailNotifConfig['useEmailNotification'] = True
        useEmailNotification = True


# Asks for the SMTP credentials
# smtplib (Not working 5/8/2022)
if useEmailNotification:
    try:
        with shelve.open('emailNotifCondig') as emailNotifCondig:
            SENDER_EMAIL = emailNotifCondig['SENDER_EMAIL']
            saveSenderEmail = emailNotifCondig['saveSenderEmail']
            askForSavePreference = emailNotifCondig['askForSavePreference']
    except KeyError:
        saveSenderEmail = False
        askForSavePreference = True

    if not saveSenderEmail:
        SENDER_EMAIL: str = pyip.inputEmail('\nPlease enter the email to be used to send the notifications. (Currently supports Gmail only. Support for Outlook and Yahoo will be added in the future. If you plan to use another email service, please notify the programmer.)\n> ')
        if askForSavePreference:
            saveSenderEmailResponse = pyip.inputYesNo('Save the email? (Whether the program will ask for your email every time it is run.)\n> ')
            if saveSenderEmailResponse == 'yes':
                with shelve.open('emailNotifCondig') as emailNotifCondig:
                    emailNotifCondig['SENDER_EMAIL'] = SENDER_EMAIL
                    emailNotifCondig['saveSenderEmail'] = True
                    emailNotifCondig['askForSavePreference']= False
            else:
                with shelve.open('emailNotifCondig') as emailNotifCondig:
                    emailNotifCondig['saveSenderEmail'] = False
                    emailNotifCondig['askForSavePreference']= False

    print(f'\nEnter the password of your email ({SENDER_EMAIL}). (The program will forget it once it is ended.)')
    if SENDER_EMAIL.endswith('@gmail.com'):
        print('If you use gmail, you will need to use an "app password." You can create one through this link https://myaccount.google.com/u/1/apppasswords\n   * Select "mail" in the "Select app" menu.')
    PASSWORD = pyip.inputPassword('> ')

    # Connects to the SMTP server
    if SENDER_EMAIL.endswith('@gmail.com'):
        while True:
            try:
                print('Logging in to the SMTP server...')
                smtpCli = smtplib.SMTP_SSL("smtp.gmail.com", 587)
                smtpCli.ehlo()
                smtpCli.starttls()
                smtpCli.login(SENDER_EMAIL, PASSWORD)
                print('Login success.')
                break
            except (TimeoutError):
                with open('log.txt', 'a', encoding='utf-8') as log:
                    log.write(traceback.format_exc())
                print('Log in failed. Please check your internet connection.')
                countdown()
                continue """
# ezgmail
def credentialsGmailInDirectory():
    for filename in os.listdir('.'):
        if filename == 'credentials.json':
            return True
    return False


useEmailNotif = True
if not credentialsGmailInDirectory():
    exitSetup = False
    useEmailNotifResponse = pyip.inputYesNo('\nDo you want to use the email notification function (Gmail)? [yes/no]\n> ')
    while not exitSetup:
        if useEmailNotifResponse == 'no':
            useEmailNotif = False
            break
        else:
            print("""Please follow these steps to be able to use Gmail:
1) Enable the Gmail API through https://console.cloud.google.com/apis/library/gmail.googleapis.com
2) Create and/or download credential file (OAuth client ID) in this link https://console.cloud.google.com/apis/api/gmail.googleapis.com/credentials
3) Rename the downloaded file (json) to 'credentials.json' and place it in the same folder as the program
   - After you continue the program, you would have to sign in and authorize the app (you only have to do this once)

NOTE: Always connect to the internet if you plan to use the email notification function.""")
            print('Press enter if done.')
            input('> ')
            exitSetup = credentialsGmailInDirectory()


if useEmailNotif:
    print('Connecting to Gmail... ', end='', flush=True)
    try:
        with shelve.open('emailNotifConfig') as emailNotifConfig:
            ezgmailInitialized = emailNotifConfig['ezgmailInitialized']
    except KeyError:
        ezgmailInitialized = False
        while not ezgmailInitialized:
            while True:
                try:
                    import ezgmail
                    ezgmail.init()
                    with shelve.open('emailNotifConfig') as emailNotifConfig:
                        emailNotifConfig['ezgmailInitialized'] = True
                    print('Connected.')
                    break
                except (UserWarning):
                    continue
            break
    if ezgmailInitialized:
        print('Connected.')

# Retrives the backup 'attendance' and 'streak' dictionaries, otherwise initializes it
with shelve.open('backup') as backup:
    try:
        attendance: Mapping[str, Mapping[str, str]] = backup['attendance']
    except KeyError:
        attendance: Mapping[str, Mapping[str, str]] = {}
with shelve.open('backup') as backup:
    try:
        streak: Mapping[str, int] = backup['streak']
    except KeyError:
        streak: Mapping[str, int] = {}


# Opens the webcam to scan QR codes
print("\nOpening the webcam... please wait\n")
cap = cv2.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

try:
    print("Webcam is running. You can now show your QR Code to the webcam.")
    print("------------------------ATTENDANCE_LOG-------------------------")

    dateCheck = ''  # 'date' cache, used to check 'date' change

    while True:
        success, img = cap.read()
        for qrcode in decode(img):
            # Appends date information to the decoded data and stores
            # it in the dictionary 'attendance'
            name = qrcode.data.decode("utf-8")
            currentTime = datetime.datetime.now()
            date = currentTime.strftime("%a %Y/%m/%d")
            clockTime = currentTime.strftime("%H:%M:%S")
            attendance.setdefault(date, {})
            if name not in attendance[date]:
                print(date, clockTime, name)
                attendance[date][name] = clockTime
                streak.setdefault(name, 0)
                streak[name] += 1

                def writeData(sheet1, name, date, dateCheck, clockTime):
                    while True:
                        try:
                            sheet1.sheet.refresh()
                            # Adds the date in the headers if the date changes while the 
                            # program is running i.e. during midnight
                            if date != dateCheck:
                                dateCheck = date
                                if not sheet1.dateInColumnHeaders(dateCheck):
                                    colNum = sheet1.findFirstBlankCol()
                                    sheet1.sheet.columnCount = colNum + 2
                                    sheet1.sheet[f'{getColumnLetterOf(colNum)}2'] = date
                                    sheet1.sheet.refresh()

                            if not sheet1.nameInRowHeaders(name):
                                rowNum = sheet1.findFirstBlankRow()
                                sheet1.sheet[f'A{rowNum}'] = name
                                sheet1.sheet.rowCount = rowNum + 2

                            row2 = sheet1.sheet.getRow(2)
                            col1 = sheet1.sheet.getColumn(1)
                            colIndex = row2.index(date) + 1
                            rowIndex = col1.index(name) + 1
                            sheet1.sheet[f'{getColumnLetterOf(colIndex)}{rowIndex}'] = clockTime
                            break
                        except (ConnectionResetError, ConnectionAbortedError, httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, TransportError, SSLEOFError):
                            print('Connection error. Please check your internet connection.')
                            silentCountdown()
                            continue
            
                # Real-time upload of data to the online database
                if useOnlineDatabase and uploadedToGoogleSheets:
                    threading.Thread(target=writeData, args=[sheet1, name, date, dateCheck, clockTime]).start()


            # Configure the design and information showed in the camera
            pts = np.array([qrcode.polygon], np.int32)
            pts = pts.reshape((-1, 1, 2))
            cv2.polylines(img, [pts], True, (255, 0, 255), 5)
            pts2 = qrcode.rect
            cv2.putText(
                img,
                name,
                (pts2[0], pts2[1]),
                cv2.FONT_HERSHEY_COMPLEX,
                0.9,
                (255, 0, 255),
                2,
            )

        cv2.imshow("Please show your QR code to the webcam.", img)
        cv2.waitKey(1)


except KeyboardInterrupt:
    # Backups the gathered data
    with shelve.open('backup') as backup:
        backup['attendance'] = attendance
        backup['streak'] = streak
    

    def getData(ws) -> Mapping[int, list]:
        rows = list(ws.rows)
        rowsDict = {}
        for i, row in enumerate(rows):
            rowsDict.setdefault(i + 1, [])
            for cell in row:
                rowsDict[i + 1].append(cell.value)

        columns = list(ws.columns)
        columnsDict = {}
        for i, column in enumerate(columns):
            columnsDict.setdefault(i + 1, [])
            for cell in column:
                columnsDict[i + 1].append(cell.value)
        return rowsDict, columnsDict


    while True:
        # Initializes the local database if not already initialized
        if noInitialFile:
            wb = openpyxl.Workbook()
            ws = wb.active
        else:
            wb = openpyxl.load_workbook(excelFilename)
            ws = wb.active
        ws.freeze_panes = 'B3'

        if ws['A1'].value != '':
            ws['A1'].value = 'Attendance'
            ws['A1'].font = Font(size=20)
        if ws['A2'].value != '':
            ws['A2'].value = 'Name'
            ws['A2'].font = Font(size=16)
        if ws['B2'].value != '':
            ws['B2'].value = 'Email'
            ws['B2'].font = Font(size=14)
        if ws['C2'].value != '':
            ws['C2'].value = 'Phone #'
            ws['C2'].font = Font(size=14)


        # Adds the necessary locators (row: name & column: date)
        rowsDict, columnsDict = getData(ws)
        for date in attendance.keys():
            if date not in rowsDict[2]:
                ws.cell(row=2, column=ws.max_column + 1).value = date
        for date in attendance:
            for name in attendance[date].keys():
                if name not in columnsDict[1]:
                    ws.cell(row=ws.max_row + 1, column=1).value = name
        
        # Stores the emails and the phone numbers to the backup to be checked
        # if there are changes in the names, emails and phone numbers
        try:
            with shelve.open('backup') as backup:
                previousEmailsAndPhoneNums: Mapping[str, Tuple[str, str]] = backup['previousEmailsAndPhoneNums']
            emailsAndPhoneNums: Mapping[str, Tuple[str, str]] = {}
        except KeyError:
            emailsAndPhoneNums: Mapping[str, Tuple[str, str]] = {}
            previousEmailsAndPhoneNums: Mapping[str, Tuple[str, str]] = {}
        for row in range(3, ws.max_row + 1):
            name = ws[f'A{row}'].value
            email = ws[f'B{row}'].value
            phone = ws[f'C{row}'].value
            emailsAndPhoneNums.setdefault(name, ('', ''))
            if emailsAndPhoneNums[name] != (email, phone):
                emailsAndPhoneNums[name] = (email, phone)

        if emailsAndPhoneNums == previousEmailsAndPhoneNums:
            noChangesInNamesEmailsAndPhoneNumbers = True
        else:
            noChangesInNamesEmailsAndPhoneNumbers = False

        with shelve.open('backup') as backup:
            backup['previousEmailsAndPhoneNums'] = emailsAndPhoneNums

        # Syncs the emails and the phone numbers to the online database.
        if useOnlineDatabase and uploadedToGoogleSheets:
            def syncEmailAndPhoneNums(emailsAndPhoneNums, sheet1: ExtendedSheet):
                while True:
                    try:
                        print('Syncing emails and phone numbers... Please wait until the sync finishes.')
                        firstBlankRow = sheet1.findFirstBlankRow()
                        col1: list = sheet1.sheet.getColumn(1)
                        for rowNum in range(2, firstBlankRow - 1):
                            name = col1[rowNum]
                            sheet1.sheet[f'B{rowNum + 1}'] = emailsAndPhoneNums[name][0]
                            sheet1.sheet[f'C{rowNum + 1}'] = emailsAndPhoneNums[name][1]
                        print('Synced successfully.\nThank you. Have a good day :)')
                        break
                    except (ConnectionResetError, ConnectionAbortedError, httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, TransportError, SSLEOFError):
                        print('Connection error. Please check your internet connection.')
                        countdown()
                        continue
            if not noChangesInNamesEmailsAndPhoneNumbers:
                syncEmailAndPhoneNumsThread = threading.Thread(target=syncEmailAndPhoneNums, args=[emailsAndPhoneNums, sheet1])
                syncEmailAndPhoneNumsThread.start()


        # Writes the gathered data to the local database if the xlsx file
        # is not currently open, otherwise restarts the initialization
        try:
            wb.save(excelFilename)
            wb = openpyxl.load_workbook(excelFilename)
            ws = wb.active
            rowsDict, columnsDict = getData(ws)
        except PermissionError:
            print("\n!!!  The excel file is open on another window  !!!")
            print("Close the excel file to let the program manipulate it. Press enter if done closing the window.")
            continueProgram = input('> ')
            continue
        else:
            for date in attendance:
                columnIndex = 0
                for column_index in columnsDict:
                    if date == columnsDict[column_index][1]:
                        columnIndex = column_index

                for name in attendance[date]:
                    rowIndex = 0
                    for row_index in rowsDict:
                        if name == rowsDict[row_index][0]:
                            rowIndex = row_index
                    ws.cell(row=rowIndex, column=columnIndex).value = attendance[date][name]

            wb.save(excelFilename)
            print(f"\nExcel file {'created' if noInitialFile else 'updated'}.\n")
            time.sleep(1)


            # Uploads the xlsx file to the online database if not already uploaded
            # (if configured to use oneline database)
            if useOnlineDatabase and not uploadedToGoogleSheets:
                while True:
                    try:
                        print('Uploading excel file to Google Sheets...')
                        ss = ezsheets.upload(excelFilename)
                    except (httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError):
                        with open('log.txt', 'a', encoding='utf-8') as log:
                            log.write(traceback.format_exc())
                        print('Upload to Google Sheets failed. Please check your internet connection.')
                        countdown(5)
                        continue
                    except (TransportError, SSLEOFError):
                        while True:
                            try:
                                uploadedToGoogleSheets, spreadsheetId = spreadsheetExists()
                                if uploadedToGoogleSheets:
                                    with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                                        googleSheetsConfig['uploadedToGoogleSheets'] = True
                                else:
                                    with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                                        googleSheetsConfig['uploadedToGoogleSheets'] = False
                                break
                            except (ConnectionAbortedError, httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, TransportError, SSLEOFError):
                                with open('log.txt', 'a', encoding='utf-8') as log:
                                    log.write(traceback.format_exc())
                                print('Update to online database failed. Please check your internet connection.')
                                countdown()
                                continue
                    else:
                        with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                            googleSheetsConfig['uploadedToGoogleSheets'] = True
                            googleSheetsConfig['spreadsheetId'] = ss.spreadsheetId
                        try:
                            with shelve.open('settings') as settings:
                                setFrozenPanes = settings['setFrozenPanes']
                        except KeyError:
                            ss['Sheet'].frozenRowCount = 2
                            ss['Sheet'].frozenColumnCount = 1
                            with shelve.open('settings') as settings:
                                settings['setFrozenPanes'] = True
                        print('Upload success.')
                        break
            break
if noChangesInNamesEmailsAndPhoneNumbers:
    print("Thank you. Have a good day :)")
