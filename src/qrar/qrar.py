try:
    import logging
    import datetime
    import shelve
    from shelve import Shelf
    import os
    import sys
    from tkinter import FALSE
    import time
    import traceback
    from contextlib import suppress
    import threading
    from typing import List, Dict, Mapping, Tuple, Union, Optional, Any
    import random
    import re
    import collections
    import smtplib
    from ssl import SSLEOFError, SSLError

    import cv2
    import numpy as np
    from pyzbar.pyzbar import decode
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils.cell import get_column_letter, column_index_from_string
    import pyinputplus
    import ezsheets
    from ezsheets import getColumnLetterOf, getColumnNumberOf
    from ezsheets import EZSheetsException
    import httplib2
    import pyinputplus as pyip
    from google.auth.exceptions import TransportError

    from quotes import strippedQuotes as quotes

    # logging.basicConfig(filename='log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
    # logging.disable(logging.DEBUG)
    logging.disable(logging.INFO)

    print(
        r"""

     o__ __o       o__ __o/               o/         o__ __o/      
    /       \     <|      \              <|         <|      \      
   />       <\    / \     <\             / \        / \     <\     
 o/           \o  \o/     o/          \o/   \o      \o/     o/     
<|__           |>  |__  _<|            |__ __|>      |__  _<|      
  \           //   |       \          /       \      |       \     
   \       \ /    <o>       \\o     o/         \\o  <o>       \\o  
    o       |      |          \    /\            \   |          \  
    <\__   /o\    / \         <\  />             <\ / \         <\ 

                 .-----------------------------.
                 | QR code Attendance Recorder |
                 '-----------------------------'
                  Programmed and maintained by:

              /\  _| _o _  _   |   _ |_  _  _ _  _ 
             /--\(_|| |(_|| )  |__(_||_)(_|_)(_|| )

| G11-Oxygen | CONTACT: zionexodus7@protonmail.com ~ 09157694749 |""")
    print("""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    ??? INSTRUCTIONS ???
1) Place the excel file at the same directory as the program (qrar.py).
2) If first used, the excel file should be completely blank i.e. no data added.
3) Press Ctrl+C to end the program.
4) If you would like to change the layout of the excel file, please notify the programmer.

    !!!   WARNING    !!!
1) Don't manipulate the excel file while the program is running.
2) Before running the program, ensure that:
    a) The excel file is closed.
    b) The excel file is not manipulated/changed in any way beforehand except by the program.
3) DO NOT DELETE the backup.* files. These files are used in the backup system and in the sync program. Delete the files only if you want to use the program for another set of data i.e. new set of students/members, new schoolyear, etc. (It won't affect the xlsx file.)
4) The googleSheetsConfig.* files contain your configuration for the online database (Google Sheets). Delete these files if you want to reset them.
5) The emailNotifConfig.* files contain your configuration for the email notification function (Gmail). Delete these files if you want to reset them.
6) The settings.* files contain the settings of the behavior of the program. Do not delete these files also.
7) DO NOT DELETE the token files. They contain your sign-in configuration for Google Sheets and Gmail.
8) If you want to delete the spreadsheet in Google Sheets (ex. if you want to reupload it), you have to delete it also in the trash folder in your google drive. If you don't, the program can still detect it and so uses it as its online database.

    ***    NOTE      ***
1) The program only records the initial time a qrcode is scanned during the day.
2) The order of the names are based on the time the qrcodes are first scanned. If you would like to sort the names alphabetically (ex. by sections), you can do so in Excel: [(Select data range by row) -> 'Data' tab -> 'SORT']. You can also do this in Google Sheets: [(Select data range by row) -> 'Data' tab -> 'Sort range'] but you don't have to since the program will automatically updates the changes in the local database to the online databse
3) The program reduces the row and column counts in the online spreadsheet to make the read/write process faster. The program will automatically increase the row and column counts as the data grows.
4) If you use the email notification function:
    a) You can edit the EDIT_EMAIL_NOTIF_MESSAGE.txt to change the subject and body of the message of the notification to be sent.
    b) You can only use the email notification function after you had added the emails of the members. However, you can only add the members' email and phone number after the xlsx file is initialized (if already created by the user or by the program).
    c) The program only sends email to those whose email is added to the xlsx file.

If you want to suggest additional functionalities or modifications to the program, feel free to contact the programmer :)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n """)

    # (See NOTE, #3)
    EXTRA_ROWS = 3
    EXTRA_COLUMNS = 3


    # Local database configuration interface
    def getExcelFiles() -> List[str]:
        excelFiles: List[str] = []
        for file in os.listdir('.'):
            if file.endswith('.xlsx'):
                excelFiles.append(file)
        return excelFiles


    while True:
        excelFiles = getExcelFiles()

        noInitialFile = False

        if len(excelFiles) == 0:
            noInitialFile = True
            print('No xlsx files found in the directory.')
            print(
                'Please enter a filename for the excel file to be created. Press enter if you would like the default name [GARSHS_ATTENDANCE.xlsx]')
            excelFilename = pyinputplus.inputStr(prompt='> ', blank=True)
            if excelFilename == '':
                excelFilename = 'GARSHS_ATTENDANCE.xlsx'
            if not excelFilename.endswith('.xlsx'):
                excelFilename += '.xlsx'
            break
        elif len(excelFiles) == 1:
            excelFilename = excelFiles[0]
            print(f'Found {excelFilename}')
            print('Exit the program (Ctrl+C) if this is not the excel file you want.')
            break
        elif len(excelFiles) == 2 and \
                (excelFiles[0].startswith('~$') or excelFiles[1].startswith('~$')):
            print('!!! The excel file is open in another window !!!')
            print('Please close the excel file then press enter to continue.')
            continueProgram = input('> ')
            continue
        else:
            exelFileIsOpen = False
            for file in os.listdir('.'):
                if exelFileIsOpen:
                    break
                if file.startswith('~$') and file.endswith('.xlsx'):
                    print('!!! An excel file is open in another window !!!')
                    print('Please close the excel file/s then press enter to continue.')
                    continueProgram = input('> ')
                    for filename in os.listdir('.'):
                        if filename.startswith('~$') and file.endswith('.xlsx'):
                            exelFileIsOpen = True
                            break
            if exelFileIsOpen:
                continue
            print('Multiple xlsx files found in the directory.')
            excelFilename: str = pyinputplus.inputMenu(getExcelFiles(), numbered=True)
            break

    # Online database configuration interface
    def countdown(seconds=3):
        print(f'Retrying in {seconds} ', end='', flush=True)
        time.sleep(1)
        count = seconds - 1
        while count > 0:
            print(f'{count} ', end='', flush=True)
            count = count - 1
            time.sleep(1)
        print()


    def silentCountdown(seconds=3):
        while seconds > 0:
            seconds -= 1
            time.sleep(1)


    def credentialsSheetsInDirectory() -> bool:
        for filename in os.listdir('.'):
            if filename == 'credentials-sheets.json':
                return True
        return False


    useOnlineDatabase = True
    if not credentialsSheetsInDirectory():
        exitSetup = False
        useOnelineDatabaseResponse = pyip.inputYesNo(
            '\nDo you want to use an online database (Google Sheets)? [yes/no]\n> ')
        while not exitSetup:
            if useOnelineDatabaseResponse == 'no':
                useOnlineDatabase = False
                break
            else:
                print("""Please follow these steps to be able to use Google Sheets:
1) Enable Google Sheets and Google Drive APIs through these links:
    * https://console.developers.google.com/apis/library/sheets.googleapis.com/
    * https://console.developers.google.com/apis/library/drive.googleapis.com/
2) Create and/or download credential file (OAuth client ID) through this link https://console.cloud.google.com/apis/credentials/oauthclient
    - You may have to configure the consent screen to create the credential file
3) Rename the downloaded file (json) to 'credentials-sheets.json' and place it in the same folder as the program
    - After you continue the program, you would have to sign in and authorize the app (twice)

NOTE: Always connect to the internet if you plan to use an online database.""")
                print('Press enter if done.')
                input('> ')
                exitSetup = credentialsSheetsInDirectory()


    # Checks if the xlsx file is already uploaded to Google Sheets
    # and if it is, gets its spreadsheetId
    # FIXME
    def spreadsheetExists():
        spreadsheets = ezsheets.listSpreadsheets()
        for key in spreadsheets:
            if spreadsheets[key] == excelFilename[:-5]:
                return True, key
        return False, None


    if useOnlineDatabase:
        try:
            with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                uploadedToGoogleSheets = googleSheetsConfig['uploadedToGoogleSheets']
        except KeyError:
            print('\nDetecting if the xlsx file is already uploaded to Google Sheets... ', end='', flush=True)
            uploadedToGoogleSheets = False
            while True:
                try:
                    uploadedToGoogleSheets, spreadsheetId = spreadsheetExists()
                    if uploadedToGoogleSheets:
                        with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                            googleSheetsConfig['uploadedToGoogleSheets'] = True
                            uploadedToGoogleSheets = True
                            googleSheetsConfig['spreadsheetId'] = spreadsheetId
                        print('The spreadsheet is found in the server.')
                    else:
                        with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                            googleSheetsConfig['uploadedToGoogleSheets'] = False
                        print(
                            'The spreadsheet is not found in the server. The workbook will be uploaded when the program ends.')
                    break
                except (ConnectionAbortedError, httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, \
                        TransportError, SSLEOFError, SSLError):
                    with open('log.txt', 'a', encoding='utf-8') as log:
                        log.write(traceback.format_exc())
                    print('Connection error. Please check your internet connection. ', end='', flush=True)
                    countdown()
                    continue


    class ExtendedSheet:
        def __init__(self, ss, title):
            self.sheet: ezsheets.Sheet = ss[title]

        def dateInColumnHeaders(self, date):
            row2 = self.sheet.getRow(2)
            for cell in row2:
                if cell == date:
                    return True, row2
            return False, row2

        def nameInRowHeaders(self, name):
            col1 = self.sheet.getColumn(1)
            for cell in col1:
                if cell == name:
                    return True, col1
            return False, col1

        @staticmethod
        def findFirstBlankCol(row2):
            colNum = 1
            while row2[colNum - 1] != '':
                colNum += 1
            return colNum

        @staticmethod
        def findFirstBlankRow(col1):
            rowNum = 1
            while col1[rowNum - 1] != '':
                rowNum += 1
            return rowNum

        @staticmethod
        def getMaxCol(row2):
            for i in range(len(row2) - 1, -1, -1):
                if row2[i] != '':
                    return i + 1

        @staticmethod
        def getMaxRow(col1):
            for i in range(len(col1) - 1, -1, -1):
                if col1[i] != '':
                    return i + 1


    if useOnlineDatabase and uploadedToGoogleSheets:
        ACTIVE_SHEET = 'Sheet'
        while True:
            with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                spreadsheetId = googleSheetsConfig['spreadsheetId']
            try:
                print('\nConnecting to Google Sheets... ', end='', flush=True)
                ss = ezsheets.Spreadsheet(spreadsheetId)
                sheet1: ExtendedSheet = ExtendedSheet(ss, ACTIVE_SHEET)
                print('Connected.')
                break
            except (
                    httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, TransportError, SSLEOFError,
                    SSLError):
                with open('log.txt', 'a', encoding='utf-8') as log:
                    log.write(traceback.format_exc())
                print('Failed to connect to Google Sheets. Please check your internet connection.')
                countdown()
                continue

    # Email notification feature configuration interface
    """ # Asks if the email notification function shall be used
    try:
        with shelve.open('emailNotifConfig') as emailNotifConfig:
            useEmailNotification = emailNotifConfig['useEmailNotification']
    except KeyError:
        useEmailNotificationResponse = pyip.inputYesNo('\nDo you want to use the email notification feature? [yes/no]\n> ')
        if useEmailNotificationResponse == 'no':
            with shelve.open('emailNotifConfig') as emailNotifConfig:
                emailNotifConfig['useEmailNotification'] = False
            useEmailNotification = False
        else:
            with shelve.open('emailNotifConfig') as emailNotifConfig:
                emailNotifConfig['useEmailNotification'] = True
            useEmailNotification = True


    # Asks for the SMTP credentials
    # smtplib (Not working 5/8/2022)
    if useEmailNotification:
        try:
            with shelve.open('emailNotifCondig') as emailNotifCondig:
                SENDER_EMAIL = emailNotifCondig['SENDER_EMAIL']
                saveSenderEmail = emailNotifCondig['saveSenderEmail']
                askForSavePreference = emailNotifCondig['askForSavePreference']
        except KeyError:
            saveSenderEmail = False
            askForSavePreference = True

        if not saveSenderEmail:
            SENDER_EMAIL: str = pyip.inputEmail('\nPlease enter the email to be used to send the notifications. (Currently supports Gmail only. Support for Outlook and Yahoo will be added in the future. If you plan to use another email service, please notify the programmer.)\n> ')
            if askForSavePreference:
                saveSenderEmailResponse = pyip.inputYesNo('Save the email? (Whether the program will ask for your email every time it is run.)\n> ')
                if saveSenderEmailResponse == 'yes':
                    with shelve.open('emailNotifCondig') as emailNotifCondig:
                        emailNotifCondig['SENDER_EMAIL'] = SENDER_EMAIL
                        emailNotifCondig['saveSenderEmail'] = True
                        emailNotifCondig['askForSavePreference']= False
                else:
                    with shelve.open('emailNotifCondig') as emailNotifCondig:
                        emailNotifCondig['saveSenderEmail'] = False
                        emailNotifCondig['askForSavePreference']= False

        print(f'\nEnter the password of your email ({SENDER_EMAIL}). (The program will forget it once it is ended.)')
        if SENDER_EMAIL.endswith('@gmail.com'):
            print('If you use gmail, you will need to use an "app password." You can create one through this link https://myaccount.google.com/u/1/apppasswords\n   * Select "mail" in the "Select app" menu.')
        PASSWORD = pyip.inputPassword('> ')

        # Connects to the SMTP server
        if SENDER_EMAIL.endswith('@gmail.com'):
            while True:
                try:
                    print('Logging in to the SMTP server...')
                    smtpCli = smtplib.SMTP_SSL("smtp.gmail.com", 587)
                    smtpCli.ehlo()
                    smtpCli.starttls()
                    smtpCli.login(SENDER_EMAIL, PASSWORD)
                    print('Login success.')
                    break
                except (TimeoutError):
                    with open('log.txt', 'a', encoding='utf-8') as log:
                        log.write(traceback.format_exc())
                    print('Log in failed. Please check your internet connection.')
                    countdown()
                    continue """


    # ezgmail
    def credentialsGmailInDirectory():
        for filename in os.listdir('.'):
            if filename == 'credentials.json':
                return True
        return False


    useEmailNotif = True
    if not credentialsGmailInDirectory():
        exitSetup = False
        useEmailNotifResponse = pyip.inputYesNo(
            '\nDo you want to use the email notification function (Gmail)? [yes/no]\n> ')
        while not exitSetup:
            if useEmailNotifResponse == 'no':
                useEmailNotif = False
                break
            else:
                print("""Please follow these steps to be able to use Gmail:
1) Enable the Gmail API through https://console.cloud.google.com/apis/library/gmail.googleapis.com
2) Create and/or download credential file (OAuth client ID) in this link https://console.cloud.google.com/apis/api/gmail.googleapis.com/credentials
3) Rename the downloaded file (json) to 'credentials.json' and place it in the same folder as the program
    - After you continue the program, you would have to sign in and authorize the app (you only have to do this once)

NOTE: Always connect to the internet if you plan to use the email notification function.""")
                print('Press enter if done.')
                input('> ')
                exitSetup = credentialsGmailInDirectory()

    # Connects to Gmail API
    if useEmailNotif:
        print('Connecting to Gmail... ', end='', flush=True)
        try:
            with shelve.open('emailNotifConfig') as emailNotifConfig:
                ezgmailInitialized = emailNotifConfig['ezgmailInitialized']
            import ezgmail
        except KeyError:
            ezgmailInitialized = False
            while not ezgmailInitialized:
                while True:
                    try:
                        import ezgmail

                        ezgmail.init()
                        with shelve.open('emailNotifConfig') as emailNotifConfig:
                            emailNotifConfig['ezgmailInitialized'] = True
                        print('Connected.')
                        break
                    except (UserWarning):
                        continue
                break
        if ezgmailInitialized:
            print('Connected.')

    # Retrieves the backup 'attendance', 'streak', and 'previousActiveCellsText' dictionaries, otherwise initializes it
    with shelve.open('backup') as backup:
        try:
            attendance: Dict[str, Dict[str, str]] = backup['attendance']
        except KeyError:
            attendance: Dict[str, Dict[str, str]] = {}
        try:
            streak: Dict[str, int] = backup['streak']
        except KeyError:
            streak: Dict[str, int] = {}
        try:
            previousActiveCellsText: List[list] = backup['previousActiveCellsText']
            logging.info(
                f"previousActiveCellsText: List[list] = backup['previousActiveCellsText'] = {backup['previousActiveCellsText']}")
        except KeyError:
            activeCellsText: List[list] = [[]]

    # Checks if there are manual changes to the worksheet
    # and if there are, syncs the changes to the online database
    wb = openpyxl.load_workbook(excelFilename)
    ws = wb.active
    activeCellsObjects = ws['A1':f'{get_column_letter(ws.max_column)}{ws.max_row}']
    activeCellsText: List[list] = [[cell.value for cell in row] for row in activeCellsObjects]
    if not noInitialFile and useOnlineDatabase and uploadedToGoogleSheets:
        logging.info(
            f"activeCellsText: List[list] = [[cell.value for cell in row] for row in activeCellsObjects] = {activeCellsText}")
        onlineDatabaseCompatibleActiveCellsText = [[' ' if cell.value is None else cell.value for cell in row] for row
                                                   in activeCellsObjects]
        if activeCellsText != previousActiveCellsText:
            print('\nDetected changes to the local database. Syncing the changes to the online database...')

            # If new rows or columns are added to the worksheet, adds the same number
            # of rows and/or columns to the previousActiveCellsText and fills them with empty strings
            if (len(activeCellsText) != len(previousActiveCellsText)) or \
                    (len(activeCellsText[0]) != len(previousActiveCellsText[0])):
                newRowsOrColumnsAddedToWorksheet = True
                newPreviousActiveCellsTextTemplate = \
                    [['' for j in range(len(activeCellsText[0]))] for i in range(len(activeCellsText))]
                for i in range(len(previousActiveCellsText)):
                    for j in range(len(previousActiveCellsText[0])):
                        newPreviousActiveCellsTextTemplate[i][j] = previousActiveCellsText[i][j]
            else:
                newRowsOrColumnsAddedToWorksheet = False

            # updateCellChangeThreads = []
            for rowNum in range(1, ws.max_row + 1):
                for colNum in range(1, ws.max_column + 1):
                    cellCoordinate = f'{get_column_letter(colNum)}{rowNum}'
                    currentCellText = activeCellsText[rowNum - 1][colNum - 1]
                    onlineDatabaseCompatibleCurrentCellText = onlineDatabaseCompatibleActiveCellsText[rowNum - 1][
                        colNum - 1]
                    if newRowsOrColumnsAddedToWorksheet:
                        worksheetValueToBeCompared = newPreviousActiveCellsTextTemplate[rowNum - 1][colNum - 1]
                    else:
                        worksheetValueToBeCompared = previousActiveCellsText[rowNum - 1][colNum - 1]
                    if currentCellText != worksheetValueToBeCompared:
                        """ def updateCellChange(sheet1, cellCoordinate, onlineDatabaseCompatibleCurrentCellText):
                            while True:
                                try:
                                    sheet1.sheet[cellCoordinate] = onlineDatabaseCompatibleCurrentCellText
                                    break
                                except (ConnectionResetError, ConnectionAbortedError, \
                                        httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, \
                                        TransportError, SSLEOFError, SSLError, AttributeError):
                                    print('Connection error. Please check your internet connection. ', end='', flush=True)
                                    countdown()
                                    continue
                        updateCellChangeThread = threading.Thread(target=updateCellChange, args=[sheet1, cellCoordinate, onlineDatabaseCompatibleCurrentCellText])
                        updateCellChangeThreads.append(updateCellChangeThread)
                        updateCellChangeThread.start()
            for updateCellThread in updateCellChangeThreads:
                updateCellThread.join() """
                        while True:
                            try:
                                sheet1.sheet[cellCoordinate] = onlineDatabaseCompatibleCurrentCellText
                                break
                            except (ConnectionResetError, ConnectionAbortedError, httplib2.ServerNotFoundError, \
                                    httplib2.error.ServerNotFoundError, TransportError, SSLEOFError, SSLError,
                                    AttributeError):
                                print('Connection error. Please check your internet connection. ', end='', flush=True)
                                countdown()
                                continue
            print('Changes synced successfully.')
        """ else:
            activeCellsText: List[list] = [[]] """

    # Put the emails and the phone numbers into a separate collection
    if useEmailNotif:
        if not ws.max_column < 2:
            emailsAndPhoneNums = collections.defaultdict(list)
            for i in range(2, len(activeCellsText)):
                emailTemp = activeCellsText[i][1] if activeCellsText[i][1] is not None else ''
                phoneTemp = activeCellsText[i][2] if activeCellsText[i][2] is not None else ''
                emailsAndPhoneNums[activeCellsText[i][0]] = [emailTemp, phoneTemp]

    # wsMaxRow, wsMaxCol = getApparentMaxRow(ws), getApparentMaxCol(ws)
    wb.save(excelFilename)
    
    # Gets the max_row and max_column of the worksheet
    # to be used later when updating the cells to the online database
    def getApparentMaxCol(ws):
        wsCols = list(ws.columns)
        colItemsNum = ws.max_row
        for i in range(len(wsCols) - 1, -1, -1):
            for j in range(colItemsNum):
                if wsCols[i][j] != '':
                    return i + 1

    def getApparentMaxRow(ws):
        wsRows = list(ws.rows)
        rowItemsNum = ws.max_column
        for i in range(len(wsRows) - 1, -1, -1):
            for j in range(rowItemsNum):2
                if wsRows[i][j] != '':
                    return i + 1

    # Opens the webcam to scan QR codes
    print("\nOpening the webcam... please wait\n")
    cap = cv2.VideoCapture(0)
    cap.set(3, 640)
    cap.set(4, 480)

    earlyExit = False
except KeyboardInterrupt:
    print('\nEnding program...')
    earlyExit = True
    with shelve.open('backup') as backup:
        try:
            attendance: Dict[str, Dict[str, str]] = backup['attendance']
        except KeyError:
            attendance: Dict[str, Dict[str, str]] = {}
        try:
            streak: Dict[str, int] = backup['streak']
        except KeyError:
            streak: Dict[str, int] = {}

try:
    if earlyExit:
        raise UserWarning
    print("Webcam is running.     You can now show your QR code to the webcam.")
    print("--------------------------ATTENDANCE_LOG---------------------------")

    # Gets the initial date to be checked with for date changes
    initialDate = datetime.datetime.now().strftime('%a %Y%m%d')

    # These will be used later when the rowCount or the columnCount have to be changed
    changedRowCount = False
    changedColumnCount = False

    while True:
        success, img = cap.read()
        for qrcode in decode(img):
            # Appends date information to the decoded data and stores
            # it in the 'attendance' dictionary
            name = qrcode.data.decode("utf-8")
            currentTime = datetime.datetime.now()
            date = currentTime.strftime("%a %Y/%m/%d")
            clockTime = currentTime.strftime("%H:%M:%S")
            attendance.setdefault(date, {})
            if name not in attendance[date]:
                print(date, clockTime, name)
                attendance[date][name] = clockTime
                streak.setdefault(name, 0)
                streak[name] += 1
                if useEmailNotif:
                    emailsAndPhoneNums.setdefault(name, [None, None])

                # Real-time upload of data to the online database
                if useOnlineDatabase and uploadedToGoogleSheets:
                    def writeData(sheet1, name, date, initialDate, clockTime):
                        global changedRowCount, changedColumnCount
                        while True:
                            try:
                                # Adds the date in the headers if the date changes while the 
                                # program is running i.e. during midnight
                                if date != initialDate:
                                    initialDate = date
                                    dateInColumnHeadersBool, row2 = sheet1.dateInColumnHeaders(date)
                                    if not dateInColumnHeadersBool:
                                        maxCol = sheet1.getMaxCol(row2)
                                        # Changes the column count of the online sheet if the added changes
                                        # in the workseet exceeds the available spaces in it
                                        # logging.info(f'wsMaxCol = {wsMaxCol}')
                                        # logging.info(f'maxCol = {maxCol}')
                                        # if wsMaxCol != maxCol and not changedColumnCount:
                                        #     # FIXME
                                        #     sheet1.sheet.columnCount = wsMaxCol + EXTRA_COLUMNS
                                        #     changedColumnCount = True

                                        # Adds the new date in the online sheet
                                        nextToMaxCol = maxCol + 1
                                        sheet1.sheet[f'{getColumnLetterOf(nextToMaxCol)}2'] = date

                                        row2 = sheet1.sheet.getRow(2)  # Gets row2 with the changes

                                # Adds name in the names column if it is new
                                nameInRowHeadersBool, col1 = sheet1.nameInRowHeaders(name)
                                if not nameInRowHeadersBool:
                                    # Changes the row count of the online sheet if the added changes
                                    # in the workseet exceeds the available spaces in it
                                    # maxRow = sheet1.getMaxRow(col1)
                                    # logging.info(f'wsMaxRow = {wsMaxRow}')
                                    # logging.info(f'maxRow = {maxRow}')
                                    # if wsMaxRow != maxRow and not changedRowCount:
                                    #     # FIXME
                                    #     sheet1.sheet.rowCount = wsMaxRow + EXTRA_ROWS
                                    #     changedRowCount = True

                                    # Adds the new name
                                    rowNum = sheet1.findFirstBlankRow(col1)
                                    sheet1.sheet[f'A{rowNum}'] = name

                                    col1 = sheet1.sheet.getColumn(1)  # Gets col1 with the changes

                                # Adds the 'clockTime' in the online sheet in the appropriate cell location
                                colIndex = row2.index(date) + 1
                                rowIndex = col1.index(name) + 1
                                sheet1.sheet[f'{getColumnLetterOf(colIndex)}{rowIndex}'] = clockTime
                                break
                            except (ConnectionResetError, ConnectionAbortedError, httplib2.ServerNotFoundError, \
                                    httplib2.error.ServerNotFoundError, TransportError, SSLEOFError, SSLError,
                                    AttributeError):
                                print('Connection error. Please check your internet connection. ', end='', flush=True)
                                countdown()
                                continue


                    threading.Thread(target=writeData, args=[sheet1, name, date, initialDate, clockTime]).start()

                # Sends the email notification only if [name]'s email is detected.
                logging.info(f'emailsAndPhoneNums = {emailsAndPhoneNums}')
                if noInitialFile:
                    email = None
                else:
                    email = emailsAndPhoneNums[name][0] if emailsAndPhoneNums[name][0] not in ('', None) else None
                logging.info(f'useEmailNotif = {useEmailNotif}')
                logging.info(f'emailsAndPhoneNums = {emailsAndPhoneNums}')
                logging.info(f'email = {email}')
                if useEmailNotif and email:
                    logging.info('Enters send mail')
                    # Gets the necessary information for the message of the notification
                    currentStreak = streak[name]
                    dailyQuote = random.choice(quotes)

                    # Extracts the subject and body part of EDIT_EMAIL_NOTIF_MESSAGE.txt
                    with open('EDIT_EMAIL_NOTIF_MESSAGE.txt') as EMAIL_NOTIF_MESSAGE:
                        lines = EMAIL_NOTIF_MESSAGE.readlines()
                    for i, line in enumerate(lines):
                        if line.strip() == 'SUBJECT:':
                            subjectLine = i
                        if line.strip() == 'BODY:':
                            bodyLine = i
                    if bodyLine - subjectLine == 2:
                        subjectRawText = lines[subjectLine + 1].strip()
                    else:
                        subjectList = []
                        for i in range(subjectLine + 1, bodyLine):
                            subjectList.append(lines[i])
                        subjectRawText = ''.join(subjectList)
                    if (len(lines) - 1) - bodyLine == 2:
                        bodyRawText = lines[bodyLine + 1].strip()
                    else:
                        bodyList = []
                        for i in range(bodyLine + 1, len(lines)):
                            bodyList.append(lines[i])
                        bodyRawText = ''.join(bodyList)


                    def substitutePlaceholders(text, clockTime, date, currentStreak, dailyQuote):
                        text = re.sub('<TIME>', clockTime, text)
                        text = re.sub('<DATE>', date, text)
                        text = re.sub('<STREAK>', str(currentStreak), text)
                        text = re.sub('<DAILYQUOTE>', dailyQuote, text)
                        return text


                    def sendEmail(name, email, date, clockTime, currentStreak, dailyQuote):
                        subject = \
                            substitutePlaceholders(subjectRawText, clockTime, date, currentStreak, dailyQuote)
                        body = \
                            substitutePlaceholders(bodyRawText, clockTime, date, currentStreak, dailyQuote)
                        while True:
                            try:
                                logging.info('Sending email...')
                                ezgmail.send(email, subject, body)
                                logging.info('Email sent.')
                                break
                            except (ConnectionResetError, ConnectionAbortedError, httplib2.ServerNotFoundError, \
                                    httplib2.error.ServerNotFoundError, TransportError, SSLEOFError, SSLError):
                                print(f'Error sending email to {email}. Please check your internet connection. ',
                                      end='', flush=True)
                                countdown()
                                continue


                    threading.Thread(target=sendEmail,
                                     args=[name, email, date, clockTime, currentStreak, dailyQuote]).start()

            # Configures the design and information showed in the camera
            pts = np.array([qrcode.polygon], np.int32)
            pts = pts.reshape((-1, 1, 2))
            cv2.polylines(img, [pts], True, (255, 0, 255), 5)
            pts2 = qrcode.rect
            cv2.putText(img, name, (pts2[0], pts2[1]), cv2.FONT_HERSHEY_COMPLEX, \
                        0.9, (255, 0, 255), 2)

        cv2.imshow("Please show your QR code to the webcam.", img)
        cv2.waitKey(1)

# (When the program is ended)
except (KeyboardInterrupt, UserWarning):
    # Backups the gathered data
    with shelve.open('backup') as backup:
        backup['attendance'] = attendance
        backup['streak'] = streak


    def getData(ws) -> Dict[int, list]:
        rows = list(ws.rows)
        rowsDict = {}
        for i, row in enumerate(rows):
            rowsDict.setdefault(i + 1, [])
            for cell in row:
                rowsDict[i + 1].append(cell.value)

        columns = list(ws.columns)
        columnsDict = {}
        for i, column in enumerate(columns):
            columnsDict.setdefault(i + 1, [])
            for cell in column:
                columnsDict[i + 1].append(cell.value)
        return rowsDict, columnsDict


    while True:
        # Initializes the local database if not already initialized
        if noInitialFile:
            wb = openpyxl.Workbook()
            ws = wb.active
        else:
            wb = openpyxl.load_workbook(excelFilename)
            ws = wb.active
        ws.freeze_panes = 'B3'

        if ws['A1'].value != '':
            ws['A1'].value = 'Attendance'
            ws['A1'].font = Font(size=20)
        if ws['A2'].value != '':
            ws['A2'].value = 'Name'
            ws['A2'].font = Font(size=16)
        if ws['B2'].value != '':
            ws['B2'].value = 'Email'
            ws['B2'].font = Font(size=14)
        if ws['C2'].value != '':
            ws['C2'].value = 'Phone #'
            ws['C2'].font = Font(size=14)


        def findFirstBlankRowInCol1(ws):
            rowNum = 3
            while True:
                currentCell = ws.cell(row=rowNum, column=1)
                if currentCell.value in (None, ''):
                    return int(currentCell.coordinate[1])
                rowNum += 1


        # Adds the necessary locators to the local databases
        # (name for the row header & date for the column header)
        rowsDict, columnsDict = getData(ws)
        for date in attendance:
            if date not in rowsDict[2]:
                ws.cell(row=2, column=ws.max_column + 1).value = date
            for name in attendance[date]:
                if name not in columnsDict[1]:
                    ws.cell(row=findFirstBlankRowInCol1(ws), column=1).value = name

        # Writes the gathered data to the local database if the xlsx file
        # is not currently open, otherwise restarts the initialization
        try:
            wb.save(excelFilename)
            wb = openpyxl.load_workbook(excelFilename)
            ws = wb.active
            rowsDict, columnsDict = getData(ws)
        except PermissionError:
            print("\n!!!  The excel file is open on another window  !!!")
            print("Close the excel file to let the program manipulate it. Press enter if done closing the window.")
            continueProgram = input('> ')
            continue
        else:
            # Adds the 'clockTimes's to the local database.
            for date in attendance:
                columnIndex = 0
                for column_index in columnsDict:
                    if date == columnsDict[column_index][1]:
                        columnIndex = column_index

                for name in attendance[date]:
                    rowIndex = 0
                    for row_index in rowsDict:
                        if name == rowsDict[row_index][0]:
                            rowIndex = row_index
                    ws.cell(row=rowIndex, column=columnIndex).value = attendance[date][name]

            wb.save(excelFilename)
            print(f"\nExcel file {'created' if noInitialFile else 'updated'}.")

            # Backups the data of the worksheet to be checked with
            # if there will be manual changes after the program ends
            wb = openpyxl.load_workbook(excelFilename)
            ws = wb.active
            activeCellsObjects = \
                ws['A1':f'{get_column_letter(ws.max_column)}{ws.max_row}']
            activeCellsText: List[list] = [[cell.value for cell in row] for row in activeCellsObjects]
            with shelve.open('backup') as backup:
                backup['previousActiveCellsText'] = activeCellsText
            logging.info(
                f"activeCellsText: List[list] = [[cell.value for cell in row] for row in activeCellsObjects]\nbackup['previousActiveCellsText'] = activeCellsText = {activeCellsText}")
            wb.save(excelFilename)

            # Uploads the xlsx file to the online database
            # if not already uploaded (if configured to use oneline database)
            if useOnlineDatabase and not uploadedToGoogleSheets:
                while True:
                    try:
                        print('Uploading excel file to Google Sheets...')
                        ss = ezsheets.upload(excelFilename)
                    except (httplib2.ServerNotFoundError, httplib2.error.ServerNotFoundError, \
                            ConnectionResetError):
                        with open('log.txt', 'a', encoding='utf-8') as log:
                            log.write(traceback.format_exc())
                        print('Upload to Google Sheets failed. Please check your internet connection. ', end='',
                              flush=True)
                        countdown(5)
                        continue
                    except (TransportError, SSLEOFError, SSLError):
                        while True:
                            try:
                                uploadedToGoogleSheets, spreadsheetId = spreadsheetExists()
                                if uploadedToGoogleSheets:
                                    with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                                        googleSheetsConfig['uploadedToGoogleSheets'] = True
                                else:
                                    with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                                        googleSheetsConfig['uploadedToGoogleSheets'] = False
                                break
                            except (ConnectionAbortedError, httplib2.ServerNotFoundError, \
                                    httplib2.error.ServerNotFoundError, TransportError, SSLEOFError, \
                                    SSLError):
                                with open('log.txt', 'a', encoding='utf-8') as log:
                                    log.write(traceback.format_exc())
                                print('Upload to online database failed. Please check your internet connection. ',
                                      end='', flush=True)
                                countdown()
                                continue
                    else:
                        with shelve.open('googleSheetsConfig') as googleSheetsConfig:
                            googleSheetsConfig['uploadedToGoogleSheets'] = True
                            googleSheetsConfig['spreadsheetId'] = ss.spreadsheetId
                        try:
                            with shelve.open('settings') as settings:
                                setFrozenPanes = settings['setFrozenPanes']
                        except KeyError:
                            ss['Sheet'].frozenRowCount = 2
                            ss['Sheet'].frozenColumnCount = 1
                            with shelve.open('settings') as settings:
                                settings['setFrozenPanes'] = True
                        print('Upload success.')
                        break
            break

print("\nThank you. Have a good day :)")
